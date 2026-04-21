"""
migrate_excel.py
────────────────
อ่านข้อมูลจาก Le Canal2026_sample.xlsx แล้วสร้างไฟล์ SQL
สำหรับ import เข้า Supabase / PostgreSQL

วิธีใช้:
  python migrate_excel.py                          # output → import_bookings.sql
  python migrate_excel.py --excel ../Le Canal2026_sample.xlsx
  python migrate_excel.py --preview                # แค่ดู ไม่เขียนไฟล์

ก่อน run: pip install openpyxl
"""

import argparse
import re
import sys
from pathlib import Path
from datetime import date

try:
    import openpyxl
except ImportError:
    print("ERROR: กรุณาติดตั้ง openpyxl ก่อน: pip install openpyxl")
    sys.exit(1)

# ─── config ────────────────────────────────────────────────────────────────

EXCEL_PATH   = Path(__file__).parent.parent / "Le Canal2026_sample.xlsx"
OUTPUT_PATH  = Path(__file__).parent / "import_bookings.sql"

# Map ชื่อ sheet → (year, month)
SHEET_MONTHS = {
    "Nov2025":  (2025, 11),
    "Dec2025":  (2025, 12),
    "Jan2026":  (2026,  1),
    "Feb2026":  (2026,  2),
    "Mar2026":  (2026,  3),
    "Apr2026":  (2026,  4),
    "May2026 ": (2026,  5),  # note trailing space
    "Jun2026":  (2026,  6),
    "July2026": (2026,  7),
    "Aug2026":  (2026,  8),
    "Sep2026":  (2026,  9),
    "Oct2026":  (2026, 10),
    "Nov2026":  (2026, 11),
    "Dec2026":  (2026, 12),
    "Jan2027":  (2027,  1),
    "Feb2027":  (2027,  2),
}

# Room name → row index ใน Excel (ดูจาก column A)
# Script จะ detect อัตโนมัติจาก column A ไม่ต้อง hardcode

# ─── helpers ───────────────────────────────────────────────────────────────

def clean_name(raw):
    """ทำความสะอาด guest name จาก cell value"""
    if raw is None or raw == 1:
        return None
    s = str(raw).strip()
    # ตัด note เวลา เช่น "Felix Luenzer  19:00 - 20:00"
    s = re.sub(r'\s+\d{1,2}:\d{2}\s*[-–]\s*\d{1,2}:\d{2}', '', s)
    # ตัด "จำนวน X ห้อง" และ note ภาษาไทย
    s = re.sub(r'\s*จำนวน.*', '', s)
    # ตัด "(X room book)" / "(X Booking)" ฯลฯ
    s = re.sub(r'\s*\([^)]*(?:room|booking|rooms)[^)]*\)', '', s, flags=re.I)
    # ตัด "X rooms" trailing
    s = re.sub(r'\s+\d+\s*rooms?', '', s, flags=re.I)
    # ตัด "/" + note หลัง slash เช่น "Day Trip Chiang Rai"
    # ไม่ตัดทิ้ง เผื่อชื่อมี slash (เช่น "Nyi Nyi/Ye")
    s = s.strip()
    return s if s else None


def sq(s):
    """Escape single-quotes for SQL"""
    return s.replace("'", "''")


# ─── parse sheet ───────────────────────────────────────────────────────────

def parse_sheet(ws, year, month):
    """
    คืนค่า list ของ dict:
      { room_col: int,  day_start: int,  day_end: int (exclusive),
        guest_name: str,  note: str }
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_row = rows[0]   # Room, Type, 1, 2, 3, …
    day_cols   = {}        # col_index → day_number
    for ci, cell in enumerate(header_row):
        if isinstance(cell, int) and 1 <= cell <= 31:
            day_cols[ci] = cell

    bookings = []

    for row in rows[2:]:   # skip header + weekday row
        room_val = row[0]
        if not isinstance(room_val, int):
            continue

        # scan columns ที่เป็น day
        pending = None   # { guest_name, note, day_start, col_start }

        sorted_cols = sorted(day_cols.keys())
        for ci in sorted_cols:
            day = day_cols[ci]
            cell_val = row[ci] if ci < len(row) else None

            is_continuation = (cell_val == 1)
            name = clean_name(cell_val) if not is_continuation else None

            if is_continuation:
                # ต่อเนื่องจาก booking ก่อนหน้า
                if pending:
                    pending['day_end'] = day + 1
                # ถ้าไม่มี pending แสดงว่า booking นี้ข้ามมาจากเดือนก่อน (ข้ามไป)

            elif name:
                # flush pending ก่อน
                if pending:
                    bookings.append({**pending, 'room_num': room_val})
                # เริ่ม booking ใหม่
                # extract note: ถ้า cell มี newline หรือ note พิเศษ
                raw_str = str(cell_val) if cell_val else ''
                note_match = re.search(r'(\d{1,2}:\d{2}.*)', raw_str)
                note = note_match.group(1).strip() if note_match else ''
                pending = {
                    'guest_name': name,
                    'note': note,
                    'day_start': day,
                    'day_end': day + 1,  # default 1 night
                }

            else:
                # empty cell → ปิด pending
                if pending:
                    bookings.append({**pending, 'room_num': room_val})
                    pending = None

        # end of row
        if pending:
            bookings.append({**pending, 'room_num': room_val})

    # แปลง day → date
    results = []
    for b in bookings:
        try:
            check_in  = date(year, month, b['day_start'])
            # check_out = check_in + nights
            # day_end เป็น exclusive day ของเดือนนั้น
            check_out_day = b['day_end']
            # ถ้า check_out_day > วันสุดท้ายของเดือน → ข้ามไปเดือนหน้า (handle แบบง่ายๆ)
            import calendar
            last_day = calendar.monthrange(year, month)[1]
            if check_out_day > last_day + 1:
                check_out_day = last_day + 1
            check_out = date(year, month, min(check_out_day, last_day))
            if check_out <= check_in:
                check_out = date(year, month, min(b['day_start'] + 1, last_day))
            results.append({
                'room_num':   b['room_num'],
                'guest_name': b['guest_name'],
                'note':       b.get('note', ''),
                'check_in':   check_in,
                'check_out':  check_out,
            })
        except ValueError:
            pass  # วันที่ invalid ข้ามไป

    return results


# ─── main ──────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='Migrate Excel bookings → SQL')
    parser.add_argument('--excel',   default=str(EXCEL_PATH), help='Path to Excel file')
    parser.add_argument('--output',  default=str(OUTPUT_PATH), help='Output SQL file')
    parser.add_argument('--preview', action='store_true', help='Print SQL instead of saving')
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"ERROR: ไม่พบไฟล์ {excel_path}")
        sys.exit(1)

    print(f"📂 อ่านไฟล์: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)

    all_bookings = []
    for sheet_name, (year, month) in SHEET_MONTHS.items():
        # หา sheet (อาจมี trailing space)
        ws = None
        for sn in wb.sheetnames:
            if sn.strip() == sheet_name.strip():
                ws = wb[sn]
                break
        if ws is None:
            print(f"  ⚠  ไม่พบ sheet '{sheet_name}' — ข้ามไป")
            continue

        bookings = parse_sheet(ws, year, month)
        print(f"  ✓  {sheet_name:12s} → {len(bookings):3d} bookings")
        all_bookings.extend(bookings)

    print(f"\n📊 รวมทั้งหมด: {len(all_bookings)} bookings\n")

    # ─── generate SQL ──────────────────────────────────────────────────────
    lines = [
        "-- Le Canal booking import",
        "-- Generated by migrate_excel.py",
        "-- Run this AFTER running schema.sql and seeding rooms",
        "--",
        "-- ⚠  ตรวจสอบ room_id ให้ตรงกับ database ของคุณก่อน!",
        "--    room_id ใน script นี้ใช้ room number ตรงๆ (1-13)",
        "--    ถ้า seed rooms ด้วย id ลำดับอื่น ต้องปรับ mapping",
        "",
        "BEGIN;",
        "",
    ]

    skipped = 0
    inserted = 0
    for b in all_bookings:
        name = b['guest_name']
        if not name:
            skipped += 1
            continue

        # ใช้ room_num เป็น room_id โดยตรง (ปรับได้ถ้า id ไม่ตรง)
        room_id = b['room_num']

        lines.append(
            f"INSERT INTO bookings (guest_name, source, room_id, check_in, check_out, status, note)"
            f" VALUES ('{sq(name)}', 'Direct', {room_id},"
            f" '{b['check_in']}', '{b['check_out']}', 'confirmed', '{sq(b['note'])}')"
            f" ON CONFLICT DO NOTHING;"
        )
        inserted += 1

    lines += [
        "",
        "COMMIT;",
        "",
        f"-- สรุป: {inserted} rows inserted, {skipped} rows skipped",
    ]

    sql = "\n".join(lines)

    if args.preview:
        print(sql[:3000])
        if len(sql) > 3000:
            print(f"\n… (แสดงแค่ 3000 chars จาก {len(sql)} chars ทั้งหมด)")
    else:
        out = Path(args.output)
        out.write_text(sql, encoding='utf-8')
        print(f"✅ เขียนไฟล์แล้ว: {out}")
        print(f"   {inserted} INSERT statements")
        print(f"\nขั้นตอนถัดไป:")
        print(f"  1. ตรวจสอบ import_bookings.sql")
        print(f"  2. Run ใน Supabase SQL Editor หรือ: psql $DATABASE_URL < import_bookings.sql")


if __name__ == '__main__':
    main()
